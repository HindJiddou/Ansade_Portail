from rest_framework import viewsets
from .models import Categorie, Theme, Tableau, Donnees,LigneIndicateur
from .serializers import CategorieSerializer, ThemeSerializer, TableauSerializer, DonneesSerializer,LigneIndicateurSerializer
import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import status, generics
import re
from collections import defaultdict,OrderedDict
from django.shortcuts import get_object_or_404
import pandas as pd
import math
from django.db.models import Q
from django.db.models import F
from rest_framework.permissions import IsAuthenticated,IsAdminUser,AllowAny
from .permissions import IsChef
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import CustomTokenObtainPairSerializer
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.authtoken.models import Token
from .serializers import UserSerializer
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from urllib.parse import unquote
from datetime import datetime
from openpyxl.utils.datetime import from_excel
import unicodedata



class CustomLoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")

        user = authenticate(request, email=email, password=password)
        if user is None:
            return Response({"error": "Email ou mot de passe incorrect"}, status=401)

        refresh = RefreshToken.for_user(user)
        user_data = UserSerializer(user).data

        return Response({
            "refresh": str(refresh),
            "access": str(refresh.access_token),
            "user": user_data
        })


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer



class UserInfoAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        return Response({
            "id": user.id,
            "email": user.email,
            "is_superuser": user.is_superuser,
            "is_chef": user.is_chef,
        })



class CategorieViewSet(viewsets.ModelViewSet):
    queryset = Categorie.objects.all()
    serializer_class = CategorieSerializer
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return []

class ThemeViewSet(viewsets.ModelViewSet):
    queryset = Theme.objects.all()
    serializer_class = ThemeSerializer
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return []

class TableauViewSet(viewsets.ModelViewSet):
    queryset = Tableau.objects.all()
    serializer_class = TableauSerializer
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsChef()]
        return []

class DonneesViewSet(viewsets.ModelViewSet):
    queryset = Donnees.objects.all()
    serializer_class = DonneesSerializer
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsChef()]
        return []


class ListeSourcesAPIView(APIView):
    def get(self, request):
        sources = Tableau.objects.exclude(source="").values_list('source', flat=True).distinct()
        return Response(sorted(sources))

def normalize_text(txt: str) -> str:
    """
    Nettoie et normalise une chaîne pour une recherche robuste.
    - Supprime les accents
    - Ignore la ponctuation
    - Met en minuscule
    """
    if not txt:
        return ""
    txt = txt.lower()
    txt = unicodedata.normalize("NFD", txt).encode("ascii", "ignore").decode("utf-8")
    txt = re.sub(r"[^\w\s]", " ", txt)  # retire ponctuation
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt

class TableauxParSourceAPIView(APIView):
    """
    Retourne les tableaux associés à une source donnée, via un paramètre GET (?source=...).
    """
    def get(self, request):
        source = request.query_params.get("source", "")
        normalized_input = normalize_text(source)

        all_tableaux = Tableau.objects.all()
        matched = [
            {"id": t.id, "titre": t.titre}
            for t in all_tableaux
            if normalized_input in normalize_text(t.source)
        ]

        return Response(matched)

import unicodedata
import re

def normalize_text(text: str) -> str:
    if not text:
        return ""
    text = text.lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return text

def contains_word_prefix(text: str, word: str) -> bool:
    """
    'age' → match 'age', 'age moyen', 'age total'
    MAIS PAS 'elevage' si mot terminé
    """
    pattern = rf"\b{re.escape(word)}"
    return re.search(pattern, text) is not None

    return re.search(pattern, text) is not None
class RechercheGlobaleAPIView(APIView):
    def get(self, request):
        raw_query = request.GET.get("q", "")
        if not raw_query.strip():
            return Response([])

        ends_with_space = raw_query.endswith(" ")
        query_norm = normalize_text(raw_query.strip())
        results = []

        def match(text_norm: str) -> bool:
            if not text_norm:
                return False

            # 🔴 CAS 1 : PAS d’espace final → recherche large
            if not ends_with_space:
                return query_norm in text_norm

            # 🟢 CAS 2 : espace final → mot terminé
            return contains_word_prefix(text_norm, query_norm)

        # Categorie
        for c in Categorie.objects.all():
            if match(normalize_text(c.nom_cat)):
                results.append({
                    "type": "Categorie",
                    "id": c.id,
                    "nom": c.nom_cat,
                })

        # Theme
        for t in Theme.objects.all():
            if match(normalize_text(t.nom_theme)):
                results.append({
                    "type": "Theme",
                    "id": t.id,
                    "nom": t.nom_theme,
                })

        # Tableau
        for tb in Tableau.objects.all():
            titre = normalize_text(tb.titre or "")
            source = normalize_text(tb.source or "")

            if match(titre) or match(source):
                results.append({
                    "type": "Tableau",
                    "id": tb.id,
                    "nom": tb.titre,
                    "source": tb.source,
                })

        return Response(results)

import re

def extraire_annee(col):
    """
    Extrait l’année d’un libellé de colonne.
    Prend la première année rencontrée : 2019, 2020, 2012, 2019/2020, etc.
    """
    match = re.search(r'(\d{4})', str(col))
    if match:
        return int(match.group(1))
    return None

def trier_tableau_par_colonnes(lignes, start_index):
    """
    Tri réel des colonnes année dans un tableau :
    - lignes = liste des lignes (liste de listes)
    - start_index = index où commencent les colonnes années
    Retourne : nouvelles_lignes (colonnes réellement déplacées)
    """

    # En-tête original
    headers = lignes[0]

    # Séparer avant + colonnes année
    colonnes_autres = list(range(start_index))
    colonnes_annee = []

    for idx in range(start_index, len(headers)):
        an = extraire_annee(headers[idx])
        if an is not None:
            colonnes_annee.append((idx, an))
        else:
            colonnes_autres.append(idx)

    # Tri des colonnes année
    colonnes_annee_tries = [idx for idx, _ in sorted(colonnes_annee, key=lambda x: x[1])]

    # Nouvel ordre complet des colonnes
    nouvel_ordre = colonnes_autres + colonnes_annee_tries

    # Reconstruire toutes les lignes
    nouvelles_lignes = []
    for row in lignes:
        new_row = []
        for idx in nouvel_ordre:
            new_row.append(row[idx] if idx < len(row) else "")
        nouvelles_lignes.append(new_row)

    return nouvelles_lignes


class ImportExcelView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsChef]

    def post(self, request):
        # Sécurité: un chef ne peut importer que dans sa catégorie
        if request.user.is_chef and not request.user.is_superuser:
            if int(request.data.get('cat_id')) != request.user.categorie_id:
                return Response({'error': 'Vous ne pouvez importer que dans votre propre catégorie'}, status=403)

        fichier_excel = request.FILES.get('file')
        id_theme = request.data.get('theme_id')
        id_cat = request.data.get('cat_id')

        if not all([fichier_excel, id_theme, id_cat]):
            return Response({'error': 'Veuillez fournir le fichier, theme_id et cat_id'}, status=400)

        try:
            wb = openpyxl.load_workbook(fichier_excel, data_only=True, read_only=True)
        except Exception as e:
            return Response({'error': f'Erreur de lecture du fichier : {str(e)}'}, status=400)

        # --------- helpers ---------
        def parse_numeric(cell):
            """Retourne (valeur_float_ou_None, had_percent_sign: bool)"""
            if isinstance(cell, str):
                had_pct = '%' in cell
                txt = (cell.replace('%', '')
                           .replace('\u202f', '')
                           .replace(' ', '')
                           .replace(',', '.')).strip()
                if txt == '':
                    return None, had_pct
                try:
                    return float(txt), had_pct
                except ValueError:
                    return None, had_pct
            elif isinstance(cell, (int, float)):
                return float(cell), False
            return None, False

        def format_excel_date(value):
            """Convertit une valeur Excel en texte mois-année (ex: janv-12)."""
            mois_fr = [
                "janv", "févr", "mars", "avr", "mai", "juin",
                "juil", "août", "sept", "oct", "nov", "déc"
            ]
            try:
                # Si Excel a stocké la date comme nombre
                if isinstance(value, (int, float)):
                    date_val = from_excel(value)
                    return f"{mois_fr[date_val.month - 1]}-{str(date_val.year)[2:]}"
                # Si c'est un datetime
                if isinstance(value, datetime):
                    return f"{mois_fr[value.month - 1]}-{str(value.year)[2:]}"
                # Si c’est une chaîne ISO comme "2012-01-01"
                if isinstance(value, str):
                    try:
                        date_val = datetime.fromisoformat(value.split(" ")[0])
                        return f"{mois_fr[date_val.month - 1]}-{str(date_val.year)[2:]}"
                    except Exception:
                        return value.strip()
                return str(value)
            except Exception:
                return str(value)

        # --------- lecture des feuilles ---------
        for feuille in wb.sheetnames:
            ws = wb[feuille]
            lignes = []
            first_non_empty_row = None

            for row in ws.iter_rows(values_only=True):
                row_str = [str(cell).strip() if cell is not None else '' for cell in row]
                if all(cell == '' for cell in row_str):
                    continue
                if first_non_empty_row is None:
                    first_non_empty_row = row_str
                lignes.append(row_str)

            if not lignes:
                continue

            normalized_first_row = [h.lower() for h in (first_non_empty_row or [])]
            is_new_format = all(col in normalized_first_row
                                for col in ["titre_fr", "source_fr", "ordre", "code", "parent", "des_fr"])

            # ==============================
            # NOUVEAU FORMAT (structuré)
            # ==============================
            if is_new_format:
                try:
                    titre_fr_idx = normalized_first_row.index('titre_fr')
                    source_fr_idx = normalized_first_row.index('source_fr')
                    ordre_idx = normalized_first_row.index('ordre')
                    code_idx = normalized_first_row.index('code')
                    parent_idx = normalized_first_row.index('parent')
                    des_fr_idx = normalized_first_row.index('des_fr')
                except ValueError as e:
                    return Response({'error': f"Colonnes manquantes : {str(e)}"}, status=400)


                # ✅ Inclure toutes les colonnes sauf 'Agreg'
                annee_indexes = []
                for i in range(des_fr_idx + 1, len(first_non_empty_row)):
                    header = str(first_non_empty_row[i]).strip()
                    if header == "" or header.lower() in ["agreg", "agrég", "agrégée"]:
                        continue  # ignorer les colonnes sans nom
                    annee_indexes.append(i)
                # ✅ Détection % dans les colonnes (NOUVEAU FORMAT)
                colonnes_pct = {}
                for idx in annee_indexes:
                    header_raw = str(first_non_empty_row[idx]).strip()
                    colonnes_pct[idx] = "%" in header_raw




                if len(lignes) < 2:
                    continue

                titre = lignes[1][titre_fr_idx]
                source = lignes[1][source_fr_idx]
                titre_lower = titre.lower()
                is_pourcentage = (
                    "%" in titre_lower
                    or "pourcentage" in titre_lower
                    or "proportion" in titre_lower
                    or "porportion" in titre_lower
                    or "taux" in titre_lower   # pour ton cas
                )


                tableau = Tableau.objects.create(
                    nom_feuille=feuille,
                    titre=titre,
                    theme_id=id_theme,
                    source=source,
                    etiquette_ligne=""
                )
                # ✅ Collecte des notes de bas de page (ex: "* Données RGE 2024")
                notes_etoiles = {}
                for row in lignes:
                    if not row:
                        continue
                    first_cell = str(row[0]).strip() if row[0] is not None else ""
                    if first_cell.startswith("*"):
                        # exemple: "* Données RGE 2024"
                        note_text = first_cell.lstrip("*").strip()
                        notes_etoiles["*"] = note_text  # on stocke la note brute


                for row in lignes[1:]:
                    label_raw = row[des_fr_idx] if len(row) > des_fr_idx else ''
                    label = format_excel_date(label_raw).strip()
                    label_raw_str = str(label_raw).strip() if label_raw else ""
                    ligne_est_pct = "%" in label_raw_str

                    # ⚠️ Si le label est vide (aucun indicateur), on saute la ligne
                    if not label:
                        continue


                    code = str(row[code_idx]).strip() if len(row) > code_idx and row[code_idx] else ''
                    parent_code = str(row[parent_idx]).strip() if len(row) > parent_idx and row[parent_idx] else ''

                    ordre = None
                    if len(row) > ordre_idx and row[ordre_idx] != '':
                        try:
                            ordre = int(float(str(row[ordre_idx]).replace(',', '.')))
                        except Exception:
                            ordre = None

                    ligne_obj = LigneIndicateur.objects.create(
                        tableau=tableau,
                        label=label,
                        code=code,
                        parent_code=parent_code,
                        ordre=ordre,
                        est_pourcentage=ligne_est_pct
                    )

                    # === Vérifie si la ligne est entièrement vide
                    ligne_vide = True

                    for idx in annee_indexes:
                        annee = format_excel_date(first_non_empty_row[idx])
                        raw_val, had_percent = (None, False)

                        # Si la colonne existe pour cette ligne
                        if idx < len(row):
                            raw_val, had_percent = parse_numeric(row[idx])

                        # Gérer la note étoilée (*)
                        note = None
                        if "*" in str(annee):
                            note = notes_etoiles.get("*", "")

                        # Si la cellule est vide
                        if raw_val is None:
                            texte_original = ""
                            if idx < len(row) and row[idx] not in [None, ""]:
                                texte_original = str(row[idx]).strip()
                                ligne_vide = False
                            else:
                                texte_original = ""

                            Donnees.objects.create(
                                ligne=ligne_obj,
                                colonne=str(annee),
                                unite="",
                                source=source,
                                valeur=None,
                                statut=texte_original,
                                categorie_id=id_cat,
                                tableau=tableau,
                                note_colonne=note
                            )
                            continue
                            
                        unite = ""
                        val = raw_val

                        is_pct_cell = (
                            ligne_est_pct                  # % dans la ligne
                            or colonnes_pct.get(idx, False) # % dans la colonne
                            or had_percent                 # % dans la cellule
                            or is_pourcentage              # % dans le titre
                        )

                        if is_pct_cell:
                            unite = "%"
                            val = round(val, 4)



                        Donnees.objects.create(
                            ligne=ligne_obj,
                            colonne=str(annee),
                            unite=unite,
                            source=source,
                            valeur=val,
                            categorie_id=id_cat,
                            tableau=tableau,
                            note_colonne=note
                        )

                    # Même si toute la ligne est vide, elle a été créée avec statut="" pour chaque colonne

                continue  # feuille traitée

            # ==============================
            # ANCIEN FORMAT
            # ==============================
            titre = ""
            source = ""
            data_start = None

            def is_non_empty(row):
                return any((str(c or "").strip() for c in row))

            def first_non_empty_after(rows, i):
                for j in range(i + 1, len(rows)):
                    if is_non_empty(rows[j]):
                        return j
                return None

            for idx, row in enumerate(lignes):
                joined = " ".join([str(c) for c in row if c]).strip()
                up = joined.upper()

                if re.search(r'\bTABLEAU\b', up) or re.match(r'^\s*TAB(?:LEAU)?\s*[\.:]?\s*\d+', up):
                    titre = joined
                    data_start = first_non_empty_after(lignes, idx)
                    continue

                    # 🔥 Détection robuste de la ligne Source
                if up.startswith("SOURCE") or up.startswith("SOURCES"):
                    src_text = re.sub(r'(?i)^sources?\s*:?', '', joined).strip()
                    if src_text:
                        source = src_text
                    continue  # 🚫 ne jamais inclure la ligne Source dans le tableau


            if not titre or data_start is None or data_start >= len(lignes):
                continue

            headers_old = lignes[data_start] if lignes[data_start] else []
            etiquette_value = str(headers_old[0]).strip() if headers_old else ""
            # ==== TRI COMPLET DES COLONNES (ANCIEN FORMAT) ====
            # On insère l'en-tête dans "lignes" pour le tri
            tableau_a_trier = [headers_old] + lignes[data_start+1:]
            tableau_trie = trier_tableau_par_colonnes(tableau_a_trier, 1)

            # Récupération
            headers_old = tableau_trie[0]
            data_rows = []

            for row in tableau_trie[1:]:
                if not row:
                    continue

                first_cell = str(row[0]).strip().lower()

                # 🛑 arrêter le tableau dès qu’on rencontre Source / Sources
                if first_cell.startswith("source") or first_cell.startswith("sources"):
                    break

                # 🛑 ignorer les notes étoilées dans les données
                if first_cell.startswith("*"):
                    break   # ou continue, mais break est plus sûr

                data_rows.append(row)


            tableau = Tableau.objects.create(
                nom_feuille=feuille,
                titre=titre,
                theme_id=id_theme,
                source=source,
                etiquette_ligne=etiquette_value
            )

            
            titre_lower = titre.lower()
            is_pourcentage = (
                "%" in titre_lower
                or "pourcentage" in titre_lower
                or "proportion" in titre_lower
                or "porportion" in titre_lower 
                or "taux" in titre_lower # pour ton cas
            )

            # ✅ Collecte des notes étoilées (ancien format)
            notes_etoiles = {}

            for row in lignes:
                if not row:
                    continue
                first_cell = str(row[0]).strip() if row[0] is not None else ""
                if first_cell.startswith("*"):
                    # Exemple: "* Données RGE 2024"
                    note_text = first_cell.lstrip("*").strip()
                    notes_etoiles["*"] = note_text
            for row in data_rows:
                # détecter la ligne source de manière robuste
                first = str(row[0]).strip().lower()
                if first.startswith("source"):
                    continue
                if first.startswith("sources"):
                    continue
                if not row or len(row) < 2:
                    continue

                # ================================
                # 🔥 DÉTECTION POURCENTAGE (ANCIEN FORMAT)
                # ================================
                indicateur_raw = (row[0] or "").strip()
                ligne_est_pct = "%" in indicateur_raw   # 👈 ICI EXACTEMENT

                indicateur_brut = format_excel_date(indicateur_raw)

                if not indicateur_brut:
                    continue

                # ================================
                # 🔥 STOCKAGE SUR LA LIGNE
                # ================================
                ligne_obj = LigneIndicateur.objects.create(
                    tableau=tableau,
                    label=indicateur_brut,
                    code='',
                    parent_code='',
                    ordre=None,
                    est_pourcentage=ligne_est_pct   # 👈 TRÈS IMPORTANT
                )


                for cidx in range(1, len(headers_old)):
                    if cidx >= len(row):
                        continue
                    header_raw = str(headers_old[cidx]).strip()
                    annee = format_excel_date(headers_old[cidx])  # ✅ colonnes
                    if not annee:
                        continue
                    # Gérer la note étoilée (*)
                    note = None
                    if "*" in str(annee):
                        note = notes_etoiles.get("*", "")


                    raw_val, had_percent = parse_numeric(row[cidx])  
                    if raw_val is None:
                        texte_original = ""
                        if cidx < len(row) and row[cidx] is not None:
                            texte_original = str(row[cidx]).strip() or ""
                       

                        # 🔥 Toujours créer la donnée (même vide)
                        Donnees.objects.create(
                            ligne=ligne_obj,
                            colonne=str(annee),
                            unite="",
                            source=source,
                            valeur=None,
                            statut=texte_original,   # <-- clé de la correction !
                            categorie_id=id_cat,
                            tableau=tableau,
                            note_colonne=note,
                        )
                        continue




                    unite = ""
                    val = raw_val
                    is_pct_colonne = "%" in header_raw

                    is_pct_cell = (
                        ligne_est_pct
                        or is_pct_colonne
                        or had_percent
                        or is_pourcentage
                    )

                    if is_pct_cell:
                        unite = "%"
                        val = round(val, 4)



                    Donnees.objects.create(
                        ligne=ligne_obj,
                        colonne=str(annee),
                        unite=unite,
                        source=source,
                        valeur=val,
                        categorie_id=id_cat,
                        tableau=tableau,
                        note_colonne=note,
                       
                    )

        return Response({'message': 'Importation réussie'}, status=201)



class TableauDetailStructureView(APIView):
    def get(self, request, tableau_id):
        donnees = (
            Donnees.objects
            .filter(tableau_id=tableau_id)
            .select_related("ligne", "tableau")
            .order_by("id")
        )

        if not donnees.exists():
            return Response({
                "colonnes_groupées": {},
                "colonnes_order": [],
                "data": [],
                "has_sous_indicateurs": False,
                "meta": {"titre": "", "source": "", "etiquette_ligne": "","afficher_decimales": tableau.afficher_decimales,
                         "tableau_id": tableau.id, "nom_feuille": tableau.nom_feuille,"theme_nom": tableau.theme.nom_theme},
                        

                "format": None,
                "statuts": [],
            })

        tableau = donnees.first().tableau
                # 🔎 Détection des colonnes pourcentage (UNE SEULE FOIS)
        colonnes_pourcentage = set()

        for d in donnees:
            col_label = (d.colonne or "").lower()
            if "%" in col_label or "pourcentage" in col_label or "taux" in col_label:
                colonnes_pourcentage.add(d.colonne)

         # ✅ Détection des statuts textuels spéciaux
        statuts_present = (
            Donnees.objects
            .filter(tableau_id=tableau_id)
            .exclude(statut__isnull=True)
            .values_list('statut', flat=True)
        )

        statuts_uniques = set([s.strip().upper() for s in statuts_present if s.strip()])
        titre_lower = (tableau.titre or "").lower()

        is_number_table = any(
            k in titre_lower
            for k in ["nombre", "effectif", "effectifs", "quantité"]
        )

        has_pct_lines = any(
            d.ligne and d.ligne.est_pourcentage
            for d in donnees
        )

        tableau_heterogene = is_number_table and has_pct_lines


        def format_value(d):
            if d.valeur is None and not d.statut:
                return ""

            if d.statut:
                return d.statut

            val = d.valeur
            ligne = d.ligne
            tableau = d.tableau

            # 🔥 PRIORITÉ ABSOLUE : % (ligne OU colonne)
            if (
                ligne.est_pourcentage
                or d.unite == "%"
                or d.colonne in colonnes_pourcentage
            ):
                return f"{val:.2f}".replace(".", ",")

            titre_lower = (tableau.titre or "").lower()
            is_number_table = any(
                k in titre_lower
                for k in ["nombre", "effectif", "effectifs", "quantité"]
            )

            # 🔢 TABLEAUX NUMÉRIQUES (purs ou hétérogènes)
            if is_number_table:
                return f"{int(round(val))}"

            # 🔘 AUTRES TABLEAUX → bouton admin
            if tableau.afficher_decimales:
                return f"{val:.2f}".replace(".", ",")

            return f"{int(round(val))}"


        # Détection format (nouveau si présence code/ordre)
        is_nouveau_format = any(
            (getattr(d.ligne, "code", None) or getattr(d.ligne, "ordre", None) is not None)
            for d in donnees
        )

        colonnes_principales = OrderedDict()

        # =====================================================================
        # NOUVEAU FORMAT (avec code/parent_code/ordre)
        # =====================================================================
        if is_nouveau_format:
            nodes_by_code = OrderedDict()

            # 1) Parcours des données -> colonnes groupées + noeuds
            for d in donnees:
                l = d.ligne
                label = (l.label or "").strip()
                col = (d.colonne or "").strip()

                # Colonnes groupées
                if "~" in col:
                    col_principal, col_sous = map(str.strip, col.split("~", 1))
                else:
                    col_principal, col_sous = col, ""
                if col_principal not in colonnes_principales:
                    colonnes_principales[col_principal] = []
                if col_sous and col_sous not in colonnes_principales[col_principal]:
                    colonnes_principales[col_principal].append(col_sous)
                elif not col_sous and "" not in colonnes_principales[col_principal]:
                    colonnes_principales[col_principal].append("")

                # Clés hiérarchie
                code = (l.code or "").strip() or f"__row_{l.id}"
                parent_code = (l.parent_code or "").strip()
                ordre = l.ordre if l.ordre is not None else None

                if code not in nodes_by_code:
                    nodes_by_code[code] = {
                        "code": code,
                        "parent_code": parent_code,
                        "indicateur": label,
                        "ordre": ordre,
                        "valeurs": defaultdict(dict),
                        "children": [],
                        "ligne_id": l.id,
                    }

                nodes_by_code[code]["valeurs"][col_principal][col_sous] = format_value(d)


            # 2) Construire l’arbre
            roots = []
            for code, node in nodes_by_code.items():
                p = node.get("parent_code")
                if p and p in nodes_by_code:
                    nodes_by_code[p]["children"].append(node)
                else:
                    roots.append(node)

            # 3) Aplatir avec tri et niveau
            def has_any_value(n):
                return any(v for g in n["valeurs"].values() for v in g.values())

            def flatten(nodes, niveau=0):
                out = []
                nodes_sorted = sorted(
                    nodes,
                    key=lambda n: (
                        n.get("ordre") is None,                # ceux sans ordre en dernier
                        n.get("ordre", 0),
                        n.get("indicateur", "")
                    )
                )
                for n in nodes_sorted:
                    out.append({
                        "indicateur": n["indicateur"],
                        "valeurs": dict(n["valeurs"]),
                        "niveau": niveau,
                        "ordre": n.get("ordre"),
                        "code": n.get("code"),
                        "parent_code": n.get("parent_code") or None,
                        "ligne_id": n.get("ligne_id"),
                        "is_section": (len(n["children"]) > 0 and not has_any_value(n)),
                    })
                    out.extend(flatten(n["children"], niveau + 1))
                return out

            data = flatten(roots)

            # 4) Colonnes groupées + ordre à plat
            colonnes_groupées = {
                col: sous if any(sous) else [""]
                for col, sous in colonnes_principales.items()
            }
            colonnes_order = []
            for gp, sous in colonnes_principales.items():
                if any(sous):
                    for s in sous:
                        colonnes_order.append({"principal": gp, "sous": s})
                else:
                    colonnes_order.append({"principal": gp, "sous": ""})

            # ✅ Récupération des notes liées aux colonnes (ex: * Données RGE 2024)
            notes = (
                Donnees.objects
                .filter(tableau_id=tableau_id)
                .exclude(note_colonne__isnull=True)
                .exclude(note_colonne__exact="")
                .values_list("note_colonne", flat=True)
                .distinct()
            )
            notes_text = [f"{n}" for n in notes]

            return Response({
                "colonnes_groupées": colonnes_groupées,
                "colonnes_order": colonnes_order,
                "data": data,
                "has_sous_indicateurs": False,
                "meta": {
                    "titre": tableau.titre,
                    "source": tableau.source or "",
                    "etiquette_ligne": tableau.etiquette_ligne or "",
                    "categorie_id": tableau.theme.categorie_id,
                    "date_verrouillage": (
                        tableau.date_verrouillage.strftime("%Y-%m-%d")
                        if tableau.date_verrouillage else None
                    ),
                    "afficher_decimales": tableau.afficher_decimales,
                    "tableau_heterogene": tableau_heterogene,
                    "tableau_numerique": is_number_table,
                    "colonnes_pourcentage": list(colonnes_pourcentage),
                    "tableau_id": tableau.id, 
                    "nom_feuille": tableau.nom_feuille,
                    "theme_nom": tableau.theme.nom_theme
                },

                "format": "nouveau",
                "notes": notes_text,  
                "statuts": list(statuts_uniques)

            })

        # =====================================================================
        # ANCIEN FORMAT (séparateur ~ dans lignes/colonnes)
        # =====================================================================
        structure = OrderedDict()

        for d in donnees:
            label = (d.ligne.label or "").strip()
            col = (d.colonne or "").strip()

            # Colonnes groupées
            if "~" in col:
                col_principal, col_sous = map(str.strip, col.split("~", 1))
            else:
                col_principal, col_sous = col, ""
            if col_principal not in colonnes_principales:
                colonnes_principales[col_principal] = []
            if col_sous and col_sous not in colonnes_principales[col_principal]:
                colonnes_principales[col_principal].append(col_sous)
            elif not col_sous and "" not in colonnes_principales[col_principal]:
                colonnes_principales[col_principal].append("")

            v = format_value(d)


            # Lignes groupées (~)
            if "~" in label:
                principal, sous = map(str.strip, label.split("~", 1))
                if principal not in structure:
                    structure[principal] = {
                        "sous_indicateurs": [],
                        "valeurs": defaultdict(dict)
                    }
                structure[principal]["sous_indicateurs"].append({
                    "nom": sous,
                    "valeurs": {col_principal: {col_sous: v}}
                })
            else:
                if label not in structure:
                    structure[label] = {
                        "sous_indicateurs": [],
                        "valeurs": defaultdict(dict)
                    }
                structure[label]["valeurs"][col_principal][col_sous] = v

        # Colonnes groupées + ordre à plat
        colonnes_groupées = {
            col: sous if any(sous) else [""]
            for col, sous in colonnes_principales.items()
        }
        colonnes_order = []
        for gp, sous in colonnes_principales.items():
            if any(sous):
                for s in sous:
                    colonnes_order.append({"principal": gp, "sous": s})
            else:
                colonnes_order.append({"principal": gp, "sous": ""})

        # Aplatir pour le front
        data = []
        for indicateur, contenu in structure.items():
            if contenu["sous_indicateurs"]:
                regroupé = OrderedDict()
                for sous in contenu["sous_indicateurs"]:
                    nom = sous["nom"]
                    if nom not in regroupé:
                        regroupé[nom] = defaultdict(dict)
                    for c, sous_vals in sous["valeurs"].items():
                        for sc, val in sous_vals.items():
                            regroupé[nom][c][sc] = val

                data.append({
                    "indicateur": indicateur,
                    "is_pourcentage": "%" in indicateur,

                    # 👉🔥 AJOUT OBLIGATOIRE POUR AFFICHER LES VALEURS PRINCIPALES
                    "valeurs": dict(contenu["valeurs"]),  # <--- ICI

                    "sous_indicateurs": [
                        {"nom": nom, "valeurs": regroupé[nom]}
                        for nom in regroupé
                    ],
                    "niveau": 0,
                    "ordre": None,
                    "code": None,
                    "parent_code": None,
                    "ligne_id": None,
                    "is_section": False
                })

            else:
                data.append({
                    "indicateur": indicateur,
                    "valeurs": dict(contenu["valeurs"]),
                    "is_pourcentage": "%" in indicateur,
                    "niveau": 0,
                    "ordre": None,
                    "code": None,
                    "parent_code": None,
                    "ligne_id": None,
                    "is_section": False
                })

        has_sous_indicateurs = any(row.get("sous_indicateurs") for row in data)

        # ✅ Récupération des notes si présentes
        notes = (
            Donnees.objects
            .filter(tableau_id=tableau_id)
            .exclude(note_colonne__isnull=True)
            .exclude(note_colonne__exact="")
            .values_list("note_colonne", flat=True)
            .distinct()
        )
        notes_text = [f"{n}" for n in notes]
       

        return Response({
            "colonnes_groupées": colonnes_groupées,
            "colonnes_order": colonnes_order,
            "data": data,
            "has_sous_indicateurs": has_sous_indicateurs,
            "meta": {
                    "titre": tableau.titre,
                    "source": tableau.source or "",
                    "etiquette_ligne": tableau.etiquette_ligne or "",
                    "categorie_id": tableau.theme.categorie_id,
                    "date_verrouillage": (
                        tableau.date_verrouillage.strftime("%Y-%m-%d")
                        if tableau.date_verrouillage else None
                    ),
                    "afficher_decimales": tableau.afficher_decimales,
                    "tableau_id": tableau.id,
                    "nom_feuille": tableau.nom_feuille,
                    "theme_nom": tableau.theme.nom_theme,
                    "tableau_heterogene": tableau_heterogene,
                    "tableau_numerique": is_number_table,
                    "colonnes_pourcentage": list(colonnes_pourcentage),


                },
            "format": "ancien",
            "notes": notes_text,  
            "statuts": list(statuts_uniques)


        })


class TableauFiltresOptionsView(APIView):
    def get(self, request, tableau_id):
        try:
            tableau = Tableau.objects.get(id=tableau_id)
        except Tableau.DoesNotExist:
            return Response({"error": "Tableau non trouvé"}, status=status.HTTP_404_NOT_FOUND)

        donnees = Donnees.objects.filter(tableau=tableau).select_related("ligne")

        lignes_set = set()
        colonnes_set = set()

        for d in donnees:
            ligne_label = d.ligne.label.strip() if d.ligne and d.ligne.label else ""
            colonne = (d.colonne or "").strip()
            if ligne_label:
                lignes_set.add(ligne_label)
            if colonne:
                colonnes_set.add(colonne)

        # Traitement des lignes
        lignes_groupées = defaultdict(list)
        for ligne in lignes_set:
            if "~" in ligne:
                indicateur, sous = map(str.strip, ligne.split("~", 1))
                lignes_groupées[indicateur].append(sous)
            else:
                lignes_groupées[ligne]

        lignes_finales = []
        for indicateur, sous_liste in lignes_groupées.items():
            if sous_liste:
                for sous in sous_liste:
                    lignes_finales.append(f"{indicateur} ~ {sous}")
            else:
                lignes_finales.append(indicateur)

        # Colonnes
        colonnes_groupées = defaultdict(list)
        for col in colonnes_set:
            if "~" in col:
                principal, sous = map(str.strip, col.split("~", 1))
                colonnes_groupées[principal].append(sous)
            else:
                colonnes_groupées[col]

        colonnes_finales = []
        for principal, sous_liste in colonnes_groupées.items():
            if sous_liste:
                for sous in sous_liste:
                    colonnes_finales.append(f"{principal} ~ {sous}")
            else:
                colonnes_finales.append(principal)

        return Response({
            "lignes": sorted(lignes_finales),
            "colonnes": sorted(colonnes_finales)
        })

class TableauFiltreView(APIView):
    def post(self, request, tableau_id):
        lignes = request.data.get("lignes", [])  # Ex: ["Population urbaine ~ Masculine"]
        colonnes = request.data.get("colonnes", [])  # Ex: ["1977", "2023"]

        try:
            tableau = Tableau.objects.get(id=tableau_id)
        except Tableau.DoesNotExist:
            return Response({"error": "Tableau non trouvé"}, status=status.HTTP_404_NOT_FOUND)
        filtres = Q(tableau=tableau)

        if lignes:
            conditions = Q()
            for ligne in lignes:
                parts = ligne.split("~")
                indicateur = parts[0].strip()
                sous = parts[1].strip() if len(parts) > 1 else None

                if sous and sous.lower() != "ensemble":
                    conditions |= Q(ligne=f"{indicateur}~{sous}")
                else:
                    conditions |= Q(ligne=indicateur)

            filtres &= conditions

        if colonnes:
            filtres &= Q(colonne__in=colonnes)

        donnees = Donnees.objects.filter(filtres)
        serializer = DonneesSerializer(donnees, many=True)
        return Response(serializer.data)



class TableauFiltreStructureView(APIView):
    def post(self, request, tableau_id):
        lignes = request.data.get("lignes", [])
        colonnes = request.data.get("colonnes", [])

        try:
            tableau = Tableau.objects.get(id=tableau_id)
        except Tableau.DoesNotExist:
            return Response({"error": "Tableau non trouvé"}, status=status.HTTP_404_NOT_FOUND)

        filtres = Q(tableau=tableau)

        # ---- Filtrage des lignes ----
        if lignes:
            conditions = Q()
            for ligne in lignes:
                parts = ligne.split("~")
                indicateur = parts[0].strip()
                sous = parts[1].strip() if len(parts) > 1 else None
                if sous and sous.lower() != "ensemble":
                    conditions |= Q(ligne__label=f"{indicateur}~{sous}")
                else:
                    conditions |= Q(ligne__label=indicateur)
            filtres &= conditions

        # ---- Filtrage des colonnes ----
        if colonnes:
            filtres &= Q(colonne__in=colonnes)

        donnees = (
            Donnees.objects.filter(filtres)
            .select_related("ligne")
            .order_by("ligne__ordre", "colonne")
        )

        if not donnees.exists():
            return Response({
                "colonnes_groupées": {},
                "data": [],
                "has_sous_indicateurs": False
            })

        # ---- Construction de la structure ----
        colonnes_principales = defaultdict(set)
        structure = defaultdict(lambda: {
            "sous_indicateurs": [],
            "valeurs": defaultdict(dict)
        })

        for d in donnees:
            ligne_label = d.ligne.label.strip() if d.ligne and d.ligne.label else ""
            colonne = (d.colonne or "").strip()
            valeur_formatee = (
                f"{d.valeur:.2f}".rstrip("0").rstrip(".") if d.valeur is not None else ""
            )
            valeur_finale = f"{valeur_formatee}{d.unite}" if d.unite else valeur_formatee

            if "~" in colonne:
                col_principal, col_sous = map(str.strip, colonne.split("~", 1))
            else:
                col_principal, col_sous = colonne, ""

            colonnes_principales[col_principal].add(col_sous)

            if "~" in ligne_label:
                principal, sous = map(str.strip, ligne_label.split("~", 1))
                structure[principal]["sous_indicateurs"].append({
                    "nom": sous,
                    "valeurs": {col_principal: {col_sous: valeur_finale}}
                })
            else:
                structure[ligne_label]["valeurs"][col_principal][col_sous] = valeur_finale

        colonnes_groupées = {
            col: sorted(list(sous)) if any(sous) else [""]
            for col, sous in colonnes_principales.items()
        }

        data = []
        for indicateur, contenu in structure.items():
            if contenu["sous_indicateurs"]:
                regroupé = defaultdict(lambda: defaultdict(dict))
                for sous in contenu["sous_indicateurs"]:
                    nom = sous["nom"]
                    for c, sous_vals in sous["valeurs"].items():
                        for sc, val in sous_vals.items():
                            regroupé[nom][c][sc] = val
                data.append({
                    "indicateur": indicateur,
                    "sous_indicateurs": [
                        {"nom": nom, "valeurs": regroupé[nom]}
                        for nom in regroupé
                    ]
                })
            else:
                data.append({
                    "indicateur": indicateur,
                    "valeurs": dict(contenu["valeurs"])
                })

        has_sous_indicateurs = any(
            row.get("sous_indicateurs") for row in structure.values()
        )

        return Response({
            "colonnes_groupées": colonnes_groupées,
            "data": data,
            "has_sous_indicateurs": has_sous_indicateurs
        })

class TableauAnalyseAPIView(APIView):
    def get(self, request, pk):
        try:
            tableau = Tableau.objects.get(id=pk)
        except Tableau.DoesNotExist:
            return Response({"detail": "Tableau non trouvé"}, status=404)

        donnees = Donnees.objects.filter(tableau=tableau)

        lignes = [d.ligne for d in donnees if d.ligne]
        colonnes = [d.colonne for d in donnees if d.colonne]

        # Détection du type de tableau
        a_wilayas = any("wilaya" in l.lower() for l in lignes)
        if all(c.isdigit() or c.startswith("20") or c.startswith("19") for c in colonnes):
            tableau_type = "annees"
        elif any(c in colonnes for c in ["Féminin", "Masculin", "Total"]):
            tableau_type = "groupes"
        elif a_wilayas:
            tableau_type = "carte"
        else:
            tableau_type = "generique"

        # Format des données
        results = []
        for d in donnees:
            try:
                valeur = float(d.valeur) if isinstance(d.valeur, str) else d.valeur
            except:
                valeur = 0
            results.append({
                "categorie_ligne": d.ligne or "",
                "categorie_colonne": d.colonne or "",
                "valeur": valeur
            })

        return Response({
            "titre": tableau.titre,
            "donnees": results,
            "type": tableau_type
        })
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Donnees, Tableau

class CarteParTableauAPIView(APIView):
    def get(self, request, tableau_id):
        try:
            tableau = Tableau.objects.get(id=tableau_id)
        except Tableau.DoesNotExist:
            return Response({'error': 'Tableau non trouvé'}, status=status.HTTP_404_NOT_FOUND)

        if tableau.etiquette_ligne.lower() != 'wilaya':
            return Response({'error': 'Ce tableau ne contient pas des données par Wilaya'}, status=status.HTTP_400_BAD_REQUEST)

        donnees = Donnees.objects.filter(tableau_id=tableau_id)

        tableau_donnees = {}

        for donnee in donnees:
            annee = donnee.colonne
            wilaya = donnee.ligne
            valeur = donnee.valeur

            if not annee or not wilaya:
                continue

            if annee not in tableau_donnees:
                tableau_donnees[annee] = {}

            if wilaya not in tableau_donnees[annee]:
                tableau_donnees[annee][wilaya] = []

            tableau_donnees[annee][wilaya].append(valeur)

        # Trie des années
        annees_triees = sorted(tableau_donnees.keys(), key=lambda x: int(x) if x.isdigit() else x)

        # Moyenne pour chaque wilaya pour la première année (la plus ancienne)
        valeurs_par_defaut = {}
        if annees_triees:
            premiere_annee = annees_triees[0]
            for wilaya, valeurs in tableau_donnees[premiere_annee].items():
                moyenne = sum(valeurs) / len(valeurs) if valeurs else 0
                valeurs_par_defaut[wilaya] = round(moyenne, 2)

        return Response({
            'titre': tableau.titre,
            'annees': annees_triees,
            'valeurs': valeurs_par_defaut
        })
    

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Tableau, Donnees
import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas


class ExportTableauAPIView(APIView):
    """
    Exporte un tableau en XLSX ou PDF.
    Exemple :
    /api/export/tableaux/5/?format=xlsx
    /api/export/tableaux/5/?format=pdf
    """

    def get(self, request, tableau_id):
        fmt = request.GET.get("format", "xlsx").lower()

        # ✅ Vérification du tableau
        tableau = get_object_or_404(Tableau, pk=tableau_id)

        # ✅ Correction : filtrer directement avec tableau_id
        donnees = Donnees.objects.filter(tableau_id=tableau_id).values()

        if not donnees.exists():
            return Response(
                {"error": "Aucune donnée disponible pour ce tableau."},
                status=status.HTTP_404_NOT_FOUND
            )

        df = pd.DataFrame(list(donnees))
        for col in ["id", "tableau_id", "categorie_id"]:
            if col in df.columns:
                df.drop(columns=[col], inplace=True)

        # === EXPORT XLSX ===
        if fmt == "xlsx":
            output = BytesIO()
            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                df.to_excel(writer, index=False, sheet_name="Données")
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            safe_name = "".join(x if x.isalnum() else "_" for x in tableau.titre)[:50]
            response["Content-Disposition"] = f'attachment; filename="{safe_name}.xlsx"'
            return response

        # === EXPORT PDF ===
        elif fmt == "pdf":
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(A4))
            p.setTitle(tableau.titre)
            p.setFont("Helvetica-Bold", 14)
            p.drawString(50, 550, f"Tableau : {tableau.titre}")
            p.setFont("Helvetica", 10)
            p.drawString(50, 530, f"Source : {tableau.source or 'ANSADE'}")

            cols = list(df.columns)
            x_start, y_start = 50, 500
            line_height = 18

            # En-têtes
            p.setFont("Helvetica-Bold", 9)
            for i, col in enumerate(cols):
                p.drawString(x_start + i * 120, y_start, str(col)[:20])

            # Lignes
            p.setFont("Helvetica", 8)
            for idx, row in enumerate(df.itertuples(index=False), start=1):
                y = y_start - idx * line_height
                if y < 40:
                    p.showPage()
                    p.setFont("Helvetica", 8)
                    y = 550
                for i, val in enumerate(row):
                    p.drawString(x_start + i * 120, y, str(val)[:20])

            p.showPage()
            p.save()
            pdf = buffer.getvalue()
            buffer.close()

            safe_name = "".join(x if x.isalnum() else "_" for x in tableau.titre)[:50]
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{safe_name}.pdf"'
            return response

        # === Format non supporté ===
        else:
            return Response(
                {"error": "Format non supporté. Utilisez ?format=pdf ou ?format=xlsx"},
                status=status.HTTP_400_BAD_REQUEST
            )

from rest_framework.views import APIView
from rest_framework.response import Response
from collections import defaultdict
import re


class GroupedSourcesAutoAPIView(APIView):
    """
    Regroupe automatiquement les sources par logique thématique
    """

    SYNONYMS = {
        "BCM": ["bcm", "banque centrale", "banque centrale de la mauritanie", "banque centrale la mauritanie"],
        "RGPH": ["rgph", "recensement général de la population"],
        "EPCV": ["epcv", "conditions de vie des ménages"],
        "MICS": ["mics", "indicateurs multiples"],
        "EDSM": ["edsm", "santé et démographie"],
        "RGE": ["rge"],
        "ICC": ["icc"],
        "PIB": ["pib"],
        "INPC": ["inpc"],
        "CN": ["cn"],
        "SNIM": ["snim"],
        "SNDE": ["snde"],
        "SOMELEC": ["somelec"],
        "SAM": ["météorologie", "sam"],
        "JUSTICE": ["prison", "justice"],
        "DOUANES": ["douanes", "douane"],
    }

    COMBOS = {
        "RGPH, MICS et EDSM": ["rgph", "mics", "edsm"],
        "RGPH et EDSM": ["rgph", "edsm"],
    }

    # ✅ ORDRE CONTRÔLÉ (au niveau de la classe)
    ORDER_ANSADE = [
        "RGPH",
        "RGPH, MICS et EDSM",
        "RGPH et EDSM",
        "EPCV",
        "MICS",
        "EDSM",
        "RGE",
        "ICC",
        "PIB",
        "INPC",
        "CN",
    ]

    ORDER_AUTRES = [
        "BCM",
        "SNIM",
        "SNDE",
        "SOMELEC",
        "SAM",
        "JUSTICE",
        "DOUANES",
    ]

    def clean_text(self, text):
        if not text:
            return ""
        text = text.lower()
        text = re.sub(r"[^\w\s]", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def detect_group(self, text):
        txt = self.clean_text(text)

        for combo_name, keywords in self.COMBOS.items():
            if all(k in txt for k in keywords):
                return combo_name

        for group, mots in self.SYNONYMS.items():
            if any(m in txt for m in mots):
                return group

        return None

    def get(self, request):
        from .models import Tableau
        from collections import defaultdict

        sources = Tableau.objects.values_list("source", flat=True).distinct()
        grouped = defaultdict(list)

        for src in sources:
            if not src:
                continue

            cleaned = src.strip()
            group = self.detect_group(cleaned)

            if group:
                grouped[group].append(cleaned)
            else:
                grouped[cleaned].append(cleaned)

        for k, v in grouped.items():
            grouped[k] = sorted(set(v))

        structure = {
            "ANSADE": {},
            "AUTRES": {},
        }

        # ✅ ANSADE : ordre contrôlé
        for key in self.ORDER_ANSADE:
            if key in grouped:
                structure["ANSADE"][key] = grouped[key]

        # ✅ AUTRES : ordre contrôlé
        for key in self.ORDER_AUTRES:
            if key in grouped:
                structure["AUTRES"][key] = grouped[key]

        # ✅ Sources restantes
        for key, srcs in grouped.items():
            if key not in self.ORDER_ANSADE and key not in self.ORDER_AUTRES:
                structure["AUTRES"][key] = srcs

        return Response(structure)


        
class TableauUpdateMetaAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, tableau_id):
        # --- Récupération du tableau ---
        try:
            tableau = Tableau.objects.get(id=tableau_id)
        except Tableau.DoesNotExist:
            return Response({"error": "Tableau non trouvé"}, status=404)

        user = request.user

        # --- 🔐 Vérification verrouillage ---
        from datetime import date
        if tableau.date_verrouillage and date.today() > tableau.date_verrouillage:
            return Response({"error": "Modification verrouillée"}, status=403)

        # --- 🔒 Permissions ---
        if user.is_superuser:
            # Superuser → autorisé pour tout
            pass

        elif user.is_chef:
            # Chef → ne peut modifier que les tableaux de SA catégorie
            if tableau.theme.categorie_id != user.categorie_id:
                return Response(
                    {"error": "Vous ne pouvez modifier que les tableaux de votre catégorie"},
                    status=403
                )

        else:
            # Autres utilisateurs → interdit
            return Response({"error": "Permission refusée"}, status=403)

        # --- Mise à jour ---
        titre = request.data.get("titre")
        source = request.data.get("source")

        if titre is not None:
            tableau.titre = titre.strip()

        if source is not None:
            tableau.source = source.strip()

        tableau.save()
        return Response({"message": "OK"})

    permission_classes = [IsAuthenticated]

    def patch(self, request, tableau_id):
        try:
            tableau = Tableau.objects.get(id=tableau_id)
        except Tableau.DoesNotExist:
            return Response({"error": "Tableau non trouvé"}, status=404)

        user = request.user

        # --- Vérification verrouillage ---
        from datetime import date
        if tableau.date_verrouillage and date.today() > tableau.date_verrouillage:
            return Response({"error": "Modification verrouillée"}, status=403)

        # --- Permissions ---
        if user.is_superuser:
            pass
        elif user.is_chef:
            if tableau.theme.categorie_id != user.categorie_id:
                return Response({"error": "Vous ne pouvez modifier que vos propres tableaux"}, status=403)
        else:
            return Response({"error": "Permission refusée"}, status=403)

        titre = request.data.get("titre")
        source = request.data.get("source")

        if titre is not None:
            tableau.titre = titre.strip()

        if source is not None:
            tableau.source = source.strip()

        tableau.save()
        return Response({"message": "OK"})

class SourceSuggestAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        term = request.query_params.get("q", "").strip().lower()
        if not term:
            return Response([])

        matches = (
            Tableau.objects
            .filter(source__icontains=term)
            .values_list("source", flat=True)
            .distinct()
        )

        return Response(list(matches))

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from .models import Theme, Tableau, Donnees


class ExportThemeAPIView(APIView):
    """
    Exporte tous les tableaux d'un thème dans un seul fichier Excel.
    Chaque tableau = une feuille Excel utilisant tableau.nom_feuille.
    """

    def get(self, request, theme_id):
        # 🔍 Vérifier le thème
        theme = get_object_or_404(Theme, pk=theme_id)

        # 🔍 Récupérer tous les tableaux du thème
        tableaux = Tableau.objects.filter(theme_id=theme_id).order_by("id")

        if not tableaux.exists():
            return Response(
                {"error": "Aucun tableau trouvé pour ce thème."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # 🟩 Créer un fichier Excel vide
        output = BytesIO()
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)   # supprimer la feuille inutile

        # 🟩 Pour chaque tableau, créer une feuille
        for tbl in tableaux:
            build_sheet_old_format(wb, tbl)

        # 🟩 Sauvegarder l'Excel
        wb.save(output)
        output.seek(0)

        # 🟩 Réponse HTTP
        safe_name = (
            f"Theme_{theme.nom_theme}"
            .replace(" ", "_")
            .replace("/", "_")
        )

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{safe_name}.xlsx"'
        return response
from openpyxl.styles import Font
from .models import LigneIndicateur, Donnees


def build_sheet_old_format(wb, tableau):

    # 1️⃣ Lignes dans l'ordre correct
    lignes = (
        LigneIndicateur.objects.filter(tableau=tableau)
        .order_by("ordre", "id")
    )

    # 2️⃣ Colonnes dans l'ordre EXACT d'apparition
    donnees_cols = (
        Donnees.objects.filter(tableau=tableau, ligne__isnull=False)
        .order_by("id")
        .values("colonne")
    )

    colonnes = []
    for d in donnees_cols:
        c = d["colonne"]
        if c not in colonnes:
            colonnes.append(c)

    # 3️⃣ Nom de la feuille (max 31 caractères)
    nom_feuille = (tableau.nom_feuille or tableau.titre)[:31]
    ws = wb.create_sheet(title=nom_feuille)

    # ----------------------------------------------
    # A) Titre
    # ----------------------------------------------
    ws.cell(row=1, column=1, value=tableau.titre).font = Font(bold=True, size=13)

    # ----------------------------------------------
    # B) En-têtes colonnes
    # ----------------------------------------------
    ws.cell(row=2, column=1, value=tableau.etiquette_ligne or "Indicateur").font = Font(bold=True)

    for i, col in enumerate(colonnes, start=2):
        ws.cell(row=2, column=i, value=col).font = Font(bold=True)

    # ----------------------------------------------
    # C) Lignes + valeurs
    # ----------------------------------------------
    row_excel = 3

    for ligne in lignes:

        ws.cell(row=row_excel, column=1, value=ligne.label)

        for ci, col_name in enumerate(colonnes, start=2):

            d = Donnees.objects.filter(
                tableau=tableau,
                ligne=ligne,
                colonne=col_name
            ).first()

            if d and d.valeur is not None:
                ws.cell(row=row_excel, column=ci, value=d.valeur)
            else:
                ws.cell(row=row_excel, column=ci, value="")

        row_excel += 1

    # ----------------------------------------------
    # D) Source (UNE SEULE FOIS)
    # ----------------------------------------------
    if tableau.source:
        ws.cell(row=row_excel + 1, column=1, value=f"Source : {tableau.source}").font = Font(italic=True)

    return ws

class TableauToggleDecimalsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, tableau_id):
        try:
            tableau = Tableau.objects.get(id=tableau_id)
        except Tableau.DoesNotExist:
            return Response({"error": "Tableau introuvable"}, status=404)

        # 🔒 Autorisation
        user = request.user
        if not user.is_superuser:
            if not (user.is_chef and user.categorie_id == tableau.theme.categorie_id):
                return Response({"error": "Non autorisé"}, status=403)

        # Nouvelle valeur reçue
        new_value = request.data.get("afficher_decimales")

        if isinstance(new_value, str):
            new_value = new_value.lower() in ["true", "1", "yes"]

        tableau.afficher_decimales = new_value
        tableau.save()

        return Response({
            "message": "Préférence mise à jour",
            "afficher_decimales": tableau.afficher_decimales
        })
